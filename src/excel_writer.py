"""
Excel Writer Module

Generates formatted Excel output files for content suggestions.
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Generate formatted Excel files for output."""
    
    def __init__(self, config: Dict):
        """
        Initialize Excel writer with configuration.
        
        Args:
            config: Configuration dictionary from config.yaml
        """
        self.config = config
        self.app_config = config.get('app', {})
        self.output_dir = Path(self.app_config.get('output_directory', 'output'))
        
        # Create output directory if it doesn't exist
        self.output_dir.mkdir(exist_ok=True)
    
    def write_existing_content_improvements(
        self,
        improvements_data: List[Dict[str, Any]],
        filename: str = "existing_content_improvements.xlsx"
    ) -> str:
        """
        Write existing content improvements to Excel.
        
        Format:
        - Column A: URL
        - Column B: Main Query/Keyword
        - Column C: Current Position
        - Column D: Current Impressions
        - Column E: Improvement Suggestions
        - Column F: Recommended Additional Keywords
        
        Args:
            improvements_data: List of improvement dictionaries
            filename: Output filename
            
        Returns:
            Path to created Excel file
        """
        logger.info(f"Writing existing content improvements to Excel...")
        
        # Prepare data for DataFrame
        rows = []
        for item in improvements_data:
            # Extract AI suggestions
            ai_suggestions = item.get('ai_suggestions', {})
            primary_improvements = ai_suggestions.get('primary_improvements', [])
            recommended_keywords = ai_suggestions.get('recommended_keywords', [])
            
            # Format improvements as bullet points
            improvements_text = "\n".join([f"• {imp}" for imp in primary_improvements])
            keywords_text = ", ".join(recommended_keywords)
            
            rows.append({
                'آدرس صفحه': item.get('url', ''),
                'کلیدواژه اصلی': item.get('main_keyword', ''),
                'موقعیت فعلی': round(item.get('position', 0), 1),
                'تعداد نمایش': item.get('impressions', 0),
                'پیشنهادات بهبود': improvements_text,
                'کلیدواژه‌های پیشنهادی': keywords_text
            })
        
        # Create DataFrame
        df = pd.DataFrame(rows)
        
        # Write to Excel
        output_path = self.output_dir / filename
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Format Excel file
        self._format_workbook(
            output_path,
            header_fill='4472C4',
            column_widths=[50, 30, 15, 18, 60, 40]
        )
        
        logger.info(f"Created: {output_path}")
        return str(output_path)
    
    def write_new_content_suggestions(
        self,
        clusters: List[Dict[str, Any]],
        filename: str = "new_content_suggestions.xlsx"
    ) -> str:
        """
        Write new content suggestions to Excel.
        
        Format:
        - Column A: Suggested Article Title
        - Column B: Predicted Impressions
        - Column C: Main Keyword Cluster
        - Columns D+: H2/H3 Headings (dynamic)
        
        Args:
            clusters: List of cluster dictionaries
            filename: Output filename
            
        Returns:
            Path to created Excel file
        """
        logger.info(f"Writing new content suggestions to Excel...")
        
        # Find maximum number of headings
        max_headings = self.app_config.get('max_headings_per_article', 8)
        max_actual_headings = max(
            [len(cluster.get('h2_headings', [])) for cluster in clusters],
            default=0
        )
        num_heading_cols = min(max_headings, max_actual_headings)
        
        # Prepare data
        rows = []
        for cluster in clusters:
            row = {
                'عنوان پیشنهادی مقاله': cluster.get('article_title', cluster.get('suggested_title', '')),
                'پیش‌بینی نمایش': int(cluster.get('avg_impressions', 0)),
                'کلاستر کلیدواژه اصلی': ", ".join(cluster.get('keywords', [])[:5]),
                'نوع محتوا': cluster.get('content_type', ''),
                'هدف جستجو': cluster.get('search_intent', ''),
                'تعداد کلمات پیشنهادی': cluster.get('recommended_word_count', 0)
            }
            
            # Add headings
            headings = cluster.get('h2_headings', [])
            for i in range(num_heading_cols):
                col_name = f'هدینگ H2 شماره {i+1}'
                row[col_name] = headings[i] if i < len(headings) else ''
            
            rows.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(rows)
        
        # Write to Excel
        output_path = self.output_dir / filename
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Format Excel file
        column_widths = [60, 18, 50, 20, 20, 20] + [40] * num_heading_cols
        self._format_workbook(
            output_path,
            header_fill='70AD47',
            column_widths=column_widths
        )
        
        logger.info(f"Created: {output_path}")
        return str(output_path)
    
    def write_full_analysis_report(
        self,
        summary_stats: Dict[str, Any],
        filename: str = "analysis_report.xlsx"
    ) -> str:
        """
        Write comprehensive analysis report with multiple sheets.
        
        Args:
            summary_stats: Dictionary with analysis statistics
            filename: Output filename
            
        Returns:
            Path to created Excel file
        """
        logger.info(f"Writing full analysis report to Excel...")
        
        output_path = self.output_dir / filename
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_df = pd.DataFrame([summary_stats])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Opportunities by position sheet
            if 'opportunities_by_position' in summary_stats:
                pos_df = pd.DataFrame(summary_stats['opportunities_by_position'])
                pos_df.to_excel(writer, sheet_name='By Position', index=False)
            
            # Top keywords sheet
            if 'top_keywords' in summary_stats:
                top_kw_df = pd.DataFrame(summary_stats['top_keywords'])
                top_kw_df.to_excel(writer, sheet_name='Top Keywords', index=False)
        
        logger.info(f"Created: {output_path}")
        return str(output_path)
    
    def _format_workbook(
        self,
        file_path: Path,
        header_fill: str = '4472C4',
        column_widths: List[int] = None
    ):
        """
        Apply formatting to Excel workbook.
        
        Args:
            file_path: Path to Excel file
            header_fill: Hex color for header background
            column_widths: List of column widths
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Header formatting
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color=header_fill, end_color=header_fill, fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Apply column widths
            if column_widths:
                for idx, width in enumerate(column_widths, start=1):
                    col_letter = get_column_letter(idx)
                    ws.column_dimensions[col_letter].width = width
            
            # Format data rows
            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(vertical='top', wrap_text=True)
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # Freeze header row
            ws.freeze_panes = ws['A2']
            
            # Auto-filter
            ws.auto_filter.ref = ws.dimensions
            
            wb.save(file_path)
            logger.info(f"Applied formatting to {file_path}")
            
        except Exception as e:
            logger.warning(f"Error formatting workbook: {str(e)}")
            # Continue even if formatting fails
    
    def write_debug_data(
        self,
        data: pd.DataFrame,
        filename: str = "debug_data.xlsx"
    ) -> str:
        """
        Write debug data to Excel for troubleshooting.
        
        Args:
            data: DataFrame to write
            filename: Output filename
            
        Returns:
            Path to created Excel file
        """
        output_path = self.output_dir / filename
        data.to_excel(output_path, index=False)
        logger.info(f"Created debug file: {output_path}")
        return str(output_path)

